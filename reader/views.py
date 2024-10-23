from django.shortcuts import render
from .forms import *
from .models import PdfReader
from django.http import HttpResponse
from io import BytesIO
import openpyxl
import pandas as pd
from copy import copy
from openpyxl.utils import range_boundaries


def reader(request):
    if request.method == "POST":
        form = LoadPdf(request.POST, request.FILES)
        if form.is_valid():
            
            pdf_file = request.FILES['pdf']
            pdf_name = pdf_file.name
            
            if pdf_file.content_type != 'application/pdf':
                return render(request, "reader/index.html", {
                    "form": form, 
                    "data": [],
                    "file_uploaded": True, 
                    "pdf_name": pdf_name,
                    "error": "El archivo cargado no es un PDF válido."
                })
            
            
            pdf_reader = PdfReader()
            result = pdf_reader.read_pdf(pdf_file)
            part_numbers = result["part_numbers"]
            last_numbers = result["last_numbers"]
            
            combined_data = list(zip(part_numbers, last_numbers))
            data = combined_data
            
            if combined_data and len(combined_data) == 0:
                data = []
            
            return render(request, "reader/index.html", {
                "form": form, 
                "data": data, 
                "file_uploaded": True, 
                "pdf_name": pdf_name 
                })
        
        elif 'excel' in request.FILES:
            excel_file = request.FILES['excel']
            wb = openpyxl.load_workbook(excel_file)
            
            values = request.POST.get('values').split(',')
            data = []
            

            for i in range(0, len(values), 2):
                part_number = values[i].strip(" ('")
                quantity = values[i + 1].strip(" ')")
                
                try:
                    part_number = pd.to_numeric(part_number)
                    quantity = pd.to_numeric(quantity)
                except ValueError:
                    part_number = None
                    quantity = None
                
                
                data.append((part_number, quantity))
            
            if len(wb.worksheets) < 2:
                error = "Excel incorrecto."
                return render(request, "reader/index.html", {"form": form, "error": error, "file_uploaded": True, "data": data})

            n = len(data)

            sheet = wb.worksheets[1]

            despiece = sheet.tables["Despiece2"]        
                
            table_range = despiece.ref
            start_cell, end_cell = table_range.split(':')
            
            _, start_row = start_cell[0], int(start_cell[1:])
            _, end_row = end_cell[0], int(end_cell[1:])
            
            new_end_row = start_row + n + 10 
            new_end_cell = f"S{new_end_row}"

            new_table_range = f"A{start_row}:{new_end_cell}"

            despiece.ref = new_table_range
            
            index = 7
            sheet.insert_rows(index, n)

            for merged in list(sheet.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merged))
                
                for row_offset in range(len(data)):
                    new_row = index + row_offset
                    if min_row <= new_row <= max_row:
                        try:
                            sheet.unmerge_cells(str(merged))
                        except KeyError as e:
                            print(f"Error unmerging cells: {e}")
                        break
            
            for row_offset in range(n):
                next_row = index + n
                new_row = index + row_offset
                
                for col in range(1, sheet.max_column + 1):
                    cell_next = sheet.cell(row=next_row, column=col)
                    cell_new = sheet.cell(row=new_row, column=col)
                    
                    if cell_next.data_type == "f":
                        cell_new.value = cell_next.value
            
                    if cell_next.has_style:
                        cell_new.font = copy(cell_next.font)
                        cell_new.border = copy(cell_next.border)
                        cell_new.fill = copy(cell_next.fill)
                        cell_new.number_format = copy(cell_next.number_format)
                        cell_new.protection = copy(cell_next.protection)
                        cell_new.alignment = copy(cell_next.alignment)


            for part_number, quantity in data:
                
                sheet[f'E{index}'] = part_number
                sheet[f'H{index}'] = quantity
                
                index += 1

            response = BytesIO()
            wb.save(response)
            response.seek(0)

            response_to_user = HttpResponse(
                response, 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response_to_user['Content-Disposition'] = 'attachment; filename=resultados_actualizados.xlsx'
            
            return response_to_user
        else:
            error = "No se cargo ningún PDF"
            return render(request, "reader/index.html", {"form": form, "error": error, "file_uploaded": False})
    else:
        form = LoadPdf()

    return render(request, "reader/index.html", {"form": form, "file_uploaded": False})